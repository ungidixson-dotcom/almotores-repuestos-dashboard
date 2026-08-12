"""
sync_comisiones_v2.py
--------------------------------------------------------------------------------
Descarga los 3 Excel de comisiones desde Dropbox via API
y los sincroniza con Supabase.

Soporta tanto refresh token (permanente) como access token (temporal).

INSTALACION (una sola vez):
  pip install openpyxl requests supabase python-dotenv

USO:
  python sync.py              # sincroniza todas las sedes
  python sync.py --sede Norte
  python sync.py --mes Julio
"""

import os, io, sys, hashlib, argparse, warnings
import requests
from datetime import datetime, date
from typing import Optional

import openpyxl
from dotenv import load_dotenv
from supabase import create_client

load_dotenv()

#  Credenciales --------------------------------------------------------
SUPABASE_URL = os.getenv('SUPABASE_URL', 'https://vvguowdjmayausyicsqs.supabase.co')
SUPABASE_KEY = os.getenv('SUPABASE_SERVICE_KEY', '')
DROPBOX_TOKEN      = os.getenv('DROPBOX_TOKEN', '')
DROPBOX_APP_KEY    = os.getenv('DROPBOX_APP_KEY', '')
DROPBOX_APP_SECRET = os.getenv('DROPBOX_APP_SECRET', '')

SEDES = {
    'Norte':     os.getenv('DROPBOX_PATH_NORTE',     '/Accesorios Norte/Comisiones Norte 2026 FO-PKIA-19 V00.xlsm'),
    'Pasoancho': os.getenv('DROPBOX_PATH_PASOANCHO', '/Accesorios 80/Comisiones Pasoancho 2026 FO-PKIA-19 V00.xlsm'),
    'Calle 9':   os.getenv('DROPBOX_PATH_CALLE9',    '/Accesorios 39/Comisiones Calle 9 2026 FO-PKIA-19 V00.xlsm'),
}

MESES_MAP = {
    'Enero':1,'Febrero':2,'Marzo':3,'Abril':4,'Mayo':5,'Junio':6,
    'Julio':7,'Agosto':8,'Septiembre':9,'Octubre':10,'Noviembre':11,'Diciembre':12,
}

COL = {
    'fecha':0,'cliente':1,'cuenta':2,'placa':3,'bastidor':4,
    'articulo':5,'denominacion':6,'cantidad':7,'neto':8,
    'prefijo_num':9,'registro':10,'asesor':11,'cedula_asesor':12,
    'comision':13,'area_venta':14,'semana':15,
}

#  Utilidades --------------------------------------------------------
def log(msg): print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")

def txt(v):
    if v is None: return None
    s = str(v).strip()
    return s if s else None

def num(v):
    try: return float(v) if v is not None else 0.0
    except: return 0.0

def fecha_iso(v):
    if v is None: return None
    if isinstance(v, (datetime, date)):
        return (v.date() if isinstance(v, datetime) else v).isoformat()
    s = str(v).strip()[:10]
    for fmt in ('%d/%m/%Y','%Y-%m-%d','%d/%m/%y'):
        try: return datetime.strptime(s, fmt).date().isoformat()
        except: pass
    return None

def fila_key(sede, anio, mes_num, registro, articulo, prefijo_num, idx=0):
    # Key determinista: sede+anio+mes+prefijo_num+articulo+registro
    # idx solo como fallback para garantizar unicidad dentro del mismo lote
    raw = f"{sede}|{anio}|{mes_num}|{prefijo_num or ''}|{articulo or ''}|{registro or ''}|{idx}"
    return hashlib.md5(raw.encode()).hexdigest()

#  Obtener access token desde refresh token 
def obtener_access_token() -> Optional[str]:
    """Intercambia el refresh token por un access token vlido por 4 horas."""
    if not DROPBOX_APP_KEY or not DROPBOX_APP_SECRET:
        return None  # usar DROPBOX_TOKEN directamente (token temporal)

    log(" Obteniendo access token desde refresh token...")
    try:
        r = requests.post(
            'https://api.dropboxapi.com/oauth2/token',
            data={
                'grant_type':    'refresh_token',
                'refresh_token': DROPBOX_TOKEN,
                'client_id':     DROPBOX_APP_KEY,
                'client_secret': DROPBOX_APP_SECRET,
            },
            timeout=30,
        )
        if r.status_code == 200:
            token = r.json().get('access_token')
            log(" Access token obtenido")
            return token
        else:
            log(f"  Error obteniendo access token: {r.text[:200]}")
            return None
    except Exception as e:
        log(f"  Error: {e}")
        return None

#  Dropbox API --------------------------------------------------------
def descargar_dropbox(path_dropbox: str, sede: str, token: str) -> Optional[io.BytesIO]:
    log(f"  Descargando {sede}...")
    try:
        r = requests.post(
            'https://content.dropboxapi.com/2/files/download',
            headers={
                'Authorization':   f'Bearer {token}',
                'Dropbox-API-Arg': f'{{"path": "{path_dropbox}"}}',
            },
            timeout=120,
        )
        if r.status_code != 200:
            log(f"  {sede}: HTTP {r.status_code}  {r.text[:200]}")
            return None
        log(f"  {sede}: {len(r.content)/1024:.0f} KB descargados")
        return io.BytesIO(r.content)
    except Exception as e:
        log(f"  {sede}: Error  {e}")
        return None

#  Lectura del Excel --------------------------------------------------------
def leer_mes(ws, sede, mes, anio_actual, filtro_mes):
    if filtro_mes and mes != filtro_mes:
        return []
    mes_num  = MESES_MAP[mes]
    registros = []
    idx = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 4: continue
        if not row or row[0] is None: continue
        f_str     = fecha_iso(row[COL['fecha']])
        anio_real = int(f_str[:4]) if f_str else anio_actual
        reg        = txt(row[COL['registro']])
        art        = txt(row[COL['articulo']])
        pref_num   = txt(row[COL['prefijo_num']])
        idx += 1
        registros.append({
            'fila_key':      fila_key(sede, anio_real, mes_num, reg, art, pref_num, idx),
            'sede':          sede,
            'anio':          anio_real,
            'mes_num':       mes_num,
            'mes':           mes,
            'semana':        txt(row[COL['semana']]),
            'fecha':         f_str,
            'cliente':       txt(row[COL['cliente']]),
            'cuenta':        txt(row[COL['cuenta']]),
            'placa':         txt(row[COL['placa']]),
            'bastidor':      txt(row[COL['bastidor']]),
            'articulo':      art,
            'denominacion':  txt(row[COL['denominacion']]),
            'cantidad':      num(row[COL['cantidad']]),
            'neto':          num(row[COL['neto']]),
            'comision':      num(row[COL['comision']]),
            'asesor':        txt(row[COL['asesor']]),
            'cedula_asesor': txt(row[COL['cedula_asesor']]),
            'area_venta':    txt(row[COL['area_venta']]),
            'prefijo_num':   pref_num,
            'registro':      reg,
            'vitrina':       sede,  # vitrina = sede por defecto desde Dropbox
            'actualizado_en': datetime.now().isoformat(),
        })
    return registros

#  Upsert Supabase --------------------------------------------------------
def limpiar_sin_vitrina(sb, sede, mes_num):
    """Elimina registros sin vitrina de un mes/sede antes de reinsertar.
    Esto evita duplicados cuando el CSV manual ya cargó datos con vitrina."""
    try:
        sb.table('comisiones_acc_detalle').delete()\
            .eq('sede', sede)\
            .eq('mes_num', mes_num)\
            .is_('vitrina', 'null')\
            .execute()
    except Exception as e:
        log(f"   Aviso limpieza previa: {e}")

def upsert(sb, registros, sede, mes):
    if not registros: return 0

    mes_num = MESES_MAP.get(mes)
    if mes_num:
        # Limpiar registros sin vitrina del mes antes de reinsertar
        # para evitar duplicados con datos cargados manualmente con vitrina
        limpiar_sin_vitrina(sb, sede, mes_num)

    total = 0
    for i in range(0, len(registros), 500):
        lote = registros[i:i+500]
        try:
            sb.table('comisiones_acc_detalle').upsert(
                lote, on_conflict='fila_key'
            ).execute()
            total += len(lote)
        except Exception as e:
            log(f"   Error lote {i//500+1}: {e}")
    return total

def log_sync(sb, sede, mes, total, error=None):
    try:
        sb.table('comisiones_acc_sync_log').insert({
            'sede': sede, 'mes': mes,
            'registros_nuevos': total, 'error': error,
        }).execute()
    except: pass

#  Main --------------------------------------------------------
def main(filtro_sede=None, filtro_mes=None):
    if not SUPABASE_KEY:
        log(" SUPABASE_SERVICE_KEY no configurada en .env"); sys.exit(1)
    if not DROPBOX_TOKEN:
        log(" DROPBOX_TOKEN no configurado en .env"); sys.exit(1)

    # Obtener access token (si hay refresh token configurado)
    token = obtener_access_token() or DROPBOX_TOKEN
    if not token:
        log(" No se pudo obtener token de Dropbox"); sys.exit(1)

    sb = create_client(SUPABASE_URL, SUPABASE_KEY)
    log(f" Conectado a Supabase")

    anio_actual = datetime.now().year
    total_global = 0

    for sede, path in SEDES.items():
        if filtro_sede and sede != filtro_sede: continue

        buf = descargar_dropbox(path, sede, token)
        if not buf:
            log_sync(sb, sede, filtro_mes or 'todos', 0, 'Error descargando')
            continue

        try:
            with warnings.catch_warnings():
                warnings.simplefilter('ignore')
                wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
        except Exception as e:
            log(f"  {sede}: Error abriendo Excel  {e}")
            log_sync(sb, sede, filtro_mes or 'todos', 0, str(e))
            continue

        for mes in MESES_MAP:
            if mes not in wb.sheetnames: continue
            registros = leer_mes(wb[mes], sede, mes, anio_actual, filtro_mes)
            if not registros:
                if not filtro_mes: log(f"    {sede}  {mes}: sin datos")
                continue
            n = upsert(sb, registros, sede, mes)
            total_global += n
            log(f"    {sede}  {mes}: {n} registros")
            log_sync(sb, sede, mes, n)

        wb.close()
        log(f"  {sede}: listo\n")

    log(f" Total: {total_global} registros sincronizados")

if __name__ == '__main__':
    p = argparse.ArgumentParser()
    p.add_argument('--sede', choices=['Norte','Pasoancho','Calle 9'])
    p.add_argument('--mes',  choices=list(MESES_MAP.keys()))
    a = p.parse_args()
    main(a.sede, a.mes)

